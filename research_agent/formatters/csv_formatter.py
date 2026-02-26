"""
CSV Formatter - Outputs structured data with consistent columns.
"""
import csv
from typing import Any, Dict, List, Optional
from pathlib import Path
from datetime import datetime

from research_agent.formatters.base import BaseFormatter


class CSVFormatter(BaseFormatter):
    """
    Formatter for CSV output.
    Best for: Product comparisons, ranked lists, tabular data
    """
    
    @property
    def file_extension(self) -> str:
        return "csv"
    
    def format(
        self, 
        data: List[Dict[str, Any]], 
        title: str,
        metadata: Optional[Dict[str, Any]] = None
    ) -> str:
        """Format data as CSV string."""
        if not data:
            return ""
        
        # Get all unique fields from all items
        all_fields = set()
        for item in data:
            all_fields.update(item.keys())
        
        # Prioritize important fields first
        priority_fields = ['rank', 'name', 'brand', 'title', 'category', 'price', 'rating', 'score']
        other_fields = sorted(f for f in all_fields if f not in priority_fields and f != 'source')
        
        # Build field order
        fields = []
        for f in priority_fields:
            if f in all_fields:
                fields.append(f)
        fields.extend(other_fields)
        if 'source' in all_fields:
            fields.append('source')
        
        # Write CSV
        import io
        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=fields, extrasaction='ignore')
        writer.writeheader()
        
        for item in data:
            # Clean up arrays for CSV display
            row = {}
            for key, value in item.items():
                if isinstance(value, list):
                    row[key] = "; ".join(str(v) for v in value)
                elif isinstance(value, dict):
                    row[key] = json.dumps(value) if 'json' in str(type(value)).lower() else str(value)
                else:
                    row[key] = value
            writer.writerow(row)
        
        return output.getvalue()
    
    def save(
        self, 
        content: str, 
        filename: str
    ) -> Path:
        """Save CSV content to file."""
        filepath = self.output_dir / f"{filename}.csv"
        filepath.write_text(content, encoding='utf-8-sig')  # UTF-8 with BOM for Excel
        return filepath


# For Excel format
import json


class ExcelFormatter(CSVFormatter):
    """
    Formatter for Excel output (.xlsx).
    Same as CSV but with .xlsx extension and Excel formatting.
    """
    
    @property
    def file_extension(self) -> str:
        return "xlsx"
    
    def save(self, content: str, filename: str) -> Path:
        """Save as Excel file using openpyxl if available, else CSV."""
        try:
            import openpyxl
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            
            # Parse CSV content
            import io
            reader = csv.DictReader(io.StringIO(content))
            rows = list(reader)
            
            if not rows:
                # Fallback to CSV
                filepath = self.output_dir / f"{filename}.csv"
                filepath.write_text(content, encoding='utf-8-sig')
                return filepath
            
            # Create Excel workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Research Results"
            
            # Headers
            headers = list(rows[0].keys())
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header.upper())
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="left", vertical="center")
            
            # Data rows
            for row_idx, row in enumerate(rows, 2):
                for col_idx, header in enumerate(headers, 1):
                    value = row.get(header, "")
                    ws.cell(row=row_idx, column=col_idx, value=value)
            
            # Auto-adjust column widths
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column].width = adjusted_width
            
            filepath = self.output_dir / f"{filename}.xlsx"
            wb.save(filepath)
            return filepath
            
        except ImportError:
            # Fallback to CSV
            filepath = self.output_dir / f"{filename}.csv"
            filepath.write_text(content, encoding='utf-8-sig')
            return filepath
